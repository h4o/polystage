from datetime import datetime

from openpyxl.styles import Alignment

from python.excel.Widgets import Widget
from openpyxl.utils import coordinate_to_tuple, get_column_letter, coordinate_from_string, column_index_from_string


class Header(Widget):
    def __init__(self, title, description):
        self.title = title
        self.description = description

    def _write(self, worksheet, cell):
        super()._write(worksheet, cell)

        self._merge_write(self.title, worksheet, cell, 0)
        self._merge_write(self.description, worksheet, cell, 1)
        self._merge_write('Generated on {}'.format(datetime.now()), worksheet, cell, 2)

    def update(self):
        pass

    def _merge_write(self, text, worksheet, base_cell, row_offset):
        row_base, col_base = coordinate_to_tuple(base_cell)
        row_cell = row_base + row_offset

        col_cell_max = col_base + self.size[0]
        cell_str = '{}{}'.format(get_column_letter(col_base), row_cell)

        worksheet.merge_cells(start_row=row_cell, start_column=col_base, end_row=row_cell, end_column=col_cell_max)
        worksheet[cell_str] = text
        worksheet[cell_str].alignment = Alignment(horizontal='center')

    @property
    def size(self):
        return 6, 3
