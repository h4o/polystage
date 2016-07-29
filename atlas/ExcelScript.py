from abc import abstractmethod

from openpyxl import Workbook

from atlas import Projects
from excel import Widgets
from util.util import pp


class ExcelScript:
    def __init__(self):
        self.wb = Workbook()
        self.nb_sheets = 0
        self.widgets = []

    def generate(self, file_name):
        self._generate()
        self._write()
        self.wb.save(file_name)

    @abstractmethod
    def _generate(self):
        pass

    def new_sheet(self, name):
        if self.nb_sheets == 0:
            ws = self.wb.active
            ws.name = name
        else:
            ws = self.wb.create_sheet(name)

        self.nb_sheets += 1
        return ws

    def put(self, widget, ws, col=1):
        self.widgets.append({
            'worksheet': ws.title,
            'col': col,
            'widget': widget
        })

    def _write(self):
        for ws in self.wb.sheetnames:
            self._write_worksheet(ws)

    def _write_worksheet(self, ws):
        cols_sizes = {}
        wid_list = [w for w in self.widgets if w['worksheet'] == ws]
        for widget in wid_list:
            widget['widget'].update()
            cols_sizes[widget['col']] = cols_sizes.get(widget['col'], [])
            cols_sizes[widget['col']].append(widget['widget'].size[0])
        offsets = [0]
        for col, val in cols_sizes.items():
            offsets.append(offsets[-1] + max(val) + 1)
        for col, offset_col in enumerate(offsets):
            offset_row = 0
            for widget in [w for w in wid_list if w['col'] == col + 1]:
                ws = self.wb[widget['worksheet']]
                widget['widget'].write(ws, 'A1', offset_col=offset_col, offset_row=offset_row)
                offset_row += widget['widget'].size[1] + 1


class IssueStats(ExcelScript):
    def __init__(self, project_tag):
        super().__init__()
        self.tag = project_tag

    def _generate(self):
        projects = Projects.GetFromTag(self.tag).do()
        for project in projects:
            ws = self.new_sheet(project['key'])
            self.put(Widgets.IssuesStatus(project['key']), ws)
            self.put(Widgets.IssuesType(project['key']), ws)
            self.put(Widgets.IssuesStatusPie(project['key']), ws, col=2)
            self.put(Widgets.AssigneePie(project['key']), ws, col=2)

            ws2 = self.new_sheet('Yayayaya')
            self.put(Widgets.IssuesStatus(project['key']), ws2)
            self.put(Widgets.IssuesType(project['key']), ws2)
            self.put(Widgets.IssuesStatusPie(project['key']), ws2, col=2)
            self.put(Widgets.AssigneePie(project['key']), ws2, col=2)
