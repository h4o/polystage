from abc import abstractmethod

from openpyxl import Workbook

from usr_scripts.excel.widgets.Header import Header


class ExcelScript:
    def __init__(self, title='', description=''):
        self.wb = Workbook()
        self.nb_sheets = 0
        self.widgets = []
        self.header = None
        if title and description:
            self.header = Header(title, description)

    def generate(self, file_name):
        self._generate()
        self._write(file_name)

    @abstractmethod
    def _generate(self):
        pass

    def new_sheet(self, name):
        if self.nb_sheets == 0:
            ws = self.wb.active
            ws.title = name
        else:
            ws = self.wb.create_sheet(name)

        self.nb_sheets += 1
        return ws

    def put(self, widget, ws, col=1):
        self.widgets.append({
            'worksheet': ws.title,
            'col': col,
            'widget': widget
        })

    def _write(self, file_name):
        for ws in self.wb.sheetnames:
            try:
                self._write_worksheet(ws)
            except Exception as e:
                print("The worksheet generation for '{}' failed : {}".format(ws, e))
            self.wb.save(file_name)

    def _write_worksheet(self, ws):
        print('Worksheet {} :'.format(ws))
        header_offset_row = 0 if self.header is None else self.header.size[1]
        if self.header:
            self.header.write(self.wb[ws], 'A1', 0, 0)

        cols_sizes = {}
        wid_list = [w for w in self.widgets if w['worksheet'] == ws]
        for widget in wid_list:
            widget['widget'].update()
            cols_sizes[widget['col']] = cols_sizes.get(widget['col'], [])
            cols_sizes[widget['col']].append(widget['widget'].size[0])
            print('  {} done'.format(widget['widget']))
        offsets = [0]
        for col, val in cols_sizes.items():
            offsets.append(offsets[-1] + max(val) + 1)
        for col, offset_col in enumerate(offsets):
            offset_row = header_offset_row + 1
            for widget in [w for w in wid_list if w['col'] == col + 1]:
                worksheet = self.wb[widget['worksheet']]
                widget['widget'].write(worksheet, 'A1', offset_col=offset_col, offset_row=offset_row)
                offset_row += widget['widget'].size[1] + 1
