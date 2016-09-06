from abc import ABC, abstractmethod

from openpyxl import chart
from openpyxl.chart import Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import WHITE, BLACK
from openpyxl.utils import coordinate_to_tuple, get_column_letter


class Widget(ABC):
    def write(self, worksheet, cell='A1', offset_col=0, offset_row=0):
        row, col = coordinate_to_tuple(cell)
        cell = get_column_letter(col + offset_col) + str(row + offset_row)

        self._write(worksheet, cell)

    @abstractmethod
    def _write(self, worksheet, cell):
        pass

    @abstractmethod
    def update(self):
        pass

    @property
    @abstractmethod
    def size(self):
        pass

    def __str__(self):
        return self.__class__.__name__


class Table(Widget):
    def __init__(self):
        self.header = []
        self.rows = []

    def _write(self, worksheet, cell):
        row, col = coordinate_to_tuple(cell)
        size_x, size_y = self.size
        rows = list(worksheet.get_squared_range(col, row, col + size_x - 1, row + size_y - 1))
        font = Font(color=WHITE, bold=True)
        head_fill = PatternFill(start_color=BLACK, end_color=BLACK, fill_type='solid')
        for i, header_cell in enumerate(rows[0]):
            header_cell.value = self.header[i]
            header_cell.font = font
            header_cell.fill = head_fill
            header_cell.alignment = Alignment(horizontal='center')
        colors = [WHITE, 'EAEAEA']
        for i, row in enumerate(rows[1:]):
            for j, c in enumerate(row):
                color = colors[i % 2]
                cell_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                c.value = self.rows[i][j]
                c.fill = cell_fill

    def append(self, *args):
        self.rows.append(args)

    @property
    def size(self):
        y = len(self.rows)
        x = 0 if y == 0 else len(self.rows[0])
        return x, y + 1


class PieChart(Table):
    def __init__(self, title):
        super().__init__()
        self.title = title
        self.header = ['key', 'value']

    def _write(self, worksheet, cell):
        super(PieChart, self)._write(worksheet, cell)
        row, col = coordinate_to_tuple(cell)

        pie = chart.PieChart()
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showVal = True
        labels = Reference(worksheet, min_col=col, min_row=row + 1, max_row=row + len(self.rows))
        data = Reference(worksheet, min_col=col + 1, min_row=row, max_row=row + len(self.rows))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = self.title
        worksheet.add_chart(pie, cell)


class LineChart(Table):
    def __init__(self, title):
        super().__init__()
        self.title = title

    def _write(self, worksheet, cell):
        super(LineChart, self)._write(worksheet, cell)
        row, col = coordinate_to_tuple(cell)

        line_chart = chart.LineChart()
        line_chart.title = self.title

        data = Reference(worksheet, min_col=col + 1, min_row=row, max_row=row + len(self.rows),
                         max_col=col + len(self.header) - 1)

        categories = Reference(worksheet, min_col=col, min_row=row + 1, max_row=row + len(self.rows))
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.x_axis.title = self.header[0]

        worksheet.add_chart(line_chart, cell)


class BarChart(Table):
    def __init__(self, title):
        super().__init__()
        self.title = title

    def _write(self, worksheet, cell):
        super(BarChart, self)._write(worksheet, cell)
        row, col = coordinate_to_tuple(cell)

        bar_chart = chart.BarChart()
        bar_chart.title = self.title

        data = Reference(worksheet, min_col=col + 1, min_row=row, max_row=row + len(self.rows),
                         max_col=col + len(self.header) - 1)

        categories = Reference(worksheet, min_col=col, min_row=row + 1, max_row=row + len(self.rows))
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        bar_chart.x_axis.title = self.header[0]

        worksheet.add_chart(bar_chart, cell)
